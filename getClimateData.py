from bs4 import BeautifulSoup
from tqdm import tqdm
import requests
import openpyxl
import time

"""
DEFINING DATA STRUCTURES
"""

class Sheet():
    def __init__(self, wb, sheet_name):
        self.wb = wb
        self.sheet_name = sheet_name
        self.sheet = None
        self.matrix_dim = None
        self.text_length = None
        self.Dates = []

    def ReadSheet(self):
        self.sheet = self.wb[self.sheet_name]

    def TextDim(self):
        sheet = self.sheet
        row_num = 0
        i = 2
        while True:
            if sheet.cell(row = i, column = 1).value is not None:
                row_num += 1
            else:
                break
            i+=1
        self.text_length = row_num

    def ReadData(self):
        for i in range(int(self.text_length)):
            actDate = self.sheet.cell(row = i + 2, column = 1).value
            dateBlock = []
            dateBlock.append(str(actDate.year))
            dateBlock.append(str(actDate.month))
            dateBlock.append(str(actDate.day))
            actDate = "-".join(dateBlock)
            self.Dates.append(actDate)

    def writeToSheet(self, RowIndex, ColumnIndex, Value):
        actCell = self.sheet.cell(row = RowIndex, column = ColumnIndex)
        actCell.value = Value

"""
DEFINING FUNCTIONS USED IN THE MAIN CYCLE
"""

def WorkBook(wb_name):
    wb = openpyxl.load_workbook(wb_name, data_only = True)
    return wb

def FtoC(F):
    return (F - 32) * 5/9

def PercentToKg(RH, T, p_abs = 101712.27):
    AH = ((6.112 * 2.7183 **((17.67 * T)/(T+243.5)) * RH * 18.02) / ((273.15 + T) * 100 * 0.08314))/1000
    airDensity = (p_abs * 0.02897) / (8.314 * (T + 273.15)) # kg/m3
    AH = AH / airDensity # kg/kg 
    return AH

def fetchClimateData(actURL, rowNum):
    try:
        req = requests.get(actURL)
        soup = BeautifulSoup(req.text, "html.parser")
        allData = soup.find_all("span", {'class': 'wu-value wu-value-to'})
        dryBulb_F = float(allData[18].next)
        dryBulb_C = round(FtoC(dryBulb_F),2)

        wetBulb_F = float(allData[21].next)
        wetBulb_C = round(FtoC(wetBulb_F),2)

        Humidity_percent = float(allData[24].next)
        Humidity_kg_kg = PercentToKg(Humidity_percent, dryBulb_C)

        return dryBulb_C, wetBulb_C, Humidity_kg_kg
    except:
        print("Error in Line {}".format(rowNum))
        return "BAD", "BAD", "BAD"


"""
MAIN CYCLE
"""

def main():
    wb = WorkBook(r"missingData.xlsx")
    sheetnames = wb.sheetnames
    actSheetName = sheetnames[0]
    sheet = Sheet(wb, actSheetName)
    sheet.ReadSheet()
    sheet.TextDim()
    sheet.ReadData()
    dates = sheet.Dates
    rowNum = 2
    for i in tqdm(range(len(dates))):
        actDate = dates[i]
        actLink = "https://www.wunderground.com/dashboard/pws/IMISKO12/graph/{}/{}/daily".format(actDate, actDate)
        dryBulb, wetBulb, Humidity = fetchClimateData(actLink, rowNum)
        sheet.writeToSheet(rowNum, 2, dryBulb)
        sheet.writeToSheet(rowNum, 3, wetBulb)
        sheet.writeToSheet(rowNum, 4, Humidity)
        rowNum += 1
    wb.save("missingData_DONE.xlsx")
main()