import wetCoolingTower as ct
import openpyxl
from tqdm import tqdm

class Sheet():
    def __init__(self, wb, sheet_name):
        self.wb = wb
        self.sheet_name = sheet_name
        
        self.sheet = None
        self.matrix_dim = None
        self.text_length = None
        self.DryCooling = None
        
        self.Dates = []
        self.DryBulb = []
        self.WetBUlb = []
        self.Humidity = []
        self.WaterFlowIn = []
        self.InletWaterTemp = []
        self.OutletWaterTemp = []

    def ReadSheet(self):
        self.sheet = self.wb[self.sheet_name]

    def TextDim(self):
        sheet = self.sheet
        row_num = 0
        i = 2
        while True:
            if sheet.cell(row = i, column = 1).value is not None:
                row_num += 1
            else:
                break
            i+=1
        self.text_length = row_num

    def ReadData(self):

        actDryCooling = self.sheet.cell(row = 2, column = 9).value
        self.DryCooling = float(actDryCooling)

        for i in range(int(self.text_length)):
            actDate = self.sheet.cell(row = i + 2, column = 1).value
            dateBlock = []
            dateBlock.append(str(actDate.year))
            dateBlock.append(str(actDate.month))
            dateBlock.append(str(actDate.day))
            actDate = "-".join(dateBlock)
            self.Dates.append(actDate)

            actDryBulb = self.sheet.cell(row = i + 2, column = 2).value
            actDryBulb = float(actDryBulb)
            self.DryBulb.append(actDryBulb)

            actHumidity = self.sheet.cell(row = i + 2, column = 4).value
            actHumidity = float(actHumidity)
            self.Humidity.append(actHumidity)

            actWaterFlowIn = self.sheet.cell(row = i + 2, column = 5).value
            actWaterFlowIn = float(actWaterFlowIn)
            self.WaterFlowIn.append(actWaterFlowIn)

            actInletWaterTemp = self.sheet.cell(row = i + 2, column = 6).value
            actWaterFlowIn = float(actInletWaterTemp)
            self.InletWaterTemp.append(actInletWaterTemp)

            actOutletWaterTemp = self.sheet.cell(row = i + 2, column = 7).value
            actOutletWaterTemp = float(actOutletWaterTemp)
            self.OutletWaterTemp.append(actOutletWaterTemp)

    def writeToSheet(self, RowIndex, ColumnIndex, Value):
        actCell = self.sheet.cell(row = RowIndex, column = ColumnIndex)
        actCell.value = Value

def WorkBook(wb_name):
    wb = openpyxl.load_workbook(wb_name, data_only = True)
    return wb


def main():
    wb = WorkBook(r"coolingTower_ORIGINAL.xlsx")
    sheetnames = wb.sheetnames
    actSheetName = sheetnames[0]
    sheet = Sheet(wb, actSheetName)

    sheet.ReadSheet()
    sheet.TextDim()
    sheet.ReadData()

    deltaT = sheet.DryCooling

    dates = sheet.Dates
    rowNum = 2

    for i in tqdm(range(len(dates))):

        try:
            WaterTempIn = sheet.InletWaterTemp[i] - deltaT
            WaterTempOut = sheet.OutletWaterTemp[i]
            AirTempInlet = sheet.DryBulb[i]
            AirHumidity = sheet.Humidity[i]
            InletWaterFlow = sheet.WaterFlowIn[i]
            Height = 2.5
            Result = ct.coolingTower(WaterTempIn, WaterTempOut, AirTempInlet, AirHumidity, InletWaterFlow, Height)
            OutletWaterFlow = Result[2]
            sheet.writeToSheet(rowNum, 8, OutletWaterFlow)
            rowNum += 1
        except:
            print("Error in line {}".format(rowNum))
            rowNum += 1
            continue
        
    wb.save(r"coolingTower_ORIGINAL_SOLVED.xlsx")
    return

main()
